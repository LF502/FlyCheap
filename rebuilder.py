from typing import Literal
from pandas import DataFrame, concat, read_csv, read_excel
from numpy import mean, nan
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import Rule
from datetime import datetime, date, timedelta
from zipfile import ZipFile
from pathlib import Path
from civilaviation import Airport, Route
from warnings import filterwarnings

class Rebuilder():
    '''
    Rebuilder
    =====
    Rebuild all data by filtering factors that influence ticket rate.
    
    Overviews
    -----
    Here are 3 significant factors can be rebuilt in class methods:
    - `airlines`: Show density and ratio by routes and hours; show flight count by airports.
    - `routes`: Show mean and std by dates and days advanced; show density and ratio by hours.
    - `dates`: Show mean rate by flight dates and collect dates (with day of week and days advanced).
    
    - Note:
        - Excels contain sheets with detailed view.
        - Time are included as a detailed view. 
        - Aircraft types are ignored for little contribution.
    
    Correlation
    -----
    Calculate Pearson correlation coefficient by modifiable groups:
    - `month`: Month of departure
    - `adv`: Days before departure
    - `week`: Day of week
    - `hour`: Hour of departure
    
    Data formatter
    -----
    Merge all rebuilt data to `DataFrame`
        - Note: Save the merged data to a csv file manually for further usage.
    
    Load data
    -----
    - `append_file`: Append an excel file.
    - `append_folder`: Append excel files from folders in `Path`.
    - `append_zip`: Append excel files from zip files in `Path`.
    - `append_data`: Append saved `DataFrame` from a `.csv` file.
    
    Parameters
    -----
    - root: `Path`, path of collection. 
    
        This should be the same for a class unless their data 
        are continuous or related.
    
    - day_limit: `int`, limit of processing days.
    
        default: `0`, no limits
    
    - starting_date: `date` | `tuple` | `int`(ordinal), the first processing date
    
        default: `0`, process all
        
    '''
    set_hyperlink = Font(u = "single", color = "0070C0")
    set_align = Alignment("center", "center")
    set_bold = Font(bold = "b")
    set_date = "mm\"-\"dd"
    set_percent = "0.00%"
    index_name = '索引-INDEX'
    cheapAir = {
        '长龙航空', '天津航', '龙江航空', '首都航', '乌航', '幸福航空', '北部湾航', 
        '西部航', '成都航空', '多彩航空', '福航', '九元航空', '金鹏航', '湖南航空', 
        '青岛航空', '长安航', '桂林航', '奥凯航空', '春秋航空', '中国联合航空', '祥鹏航空'
    }
    
    def __init__(self, root: Path | str = Path(), 
                 starting_date: date | tuple | int = 0, day_limit: int = 0) -> None:
        
        self.__merge = DataFrame()
        self.__preprocess = DataFrame()
        
        self.__header_min = {
            'date_flight', 'day_week', 'airline', 'type', 'dep',
            'arr', 'time_dep', 'time_arr', 'price', 'price_rate'
        }
        self.__header_req = {'day_adv', 'hour_dep', 'route'}
        self.__title = {
            'date_flight': '航班日期', 'day_week': '星期', 
            'date_coll': '收集日期', 'day_adv': '提前天数', 
            'airline': '航司', 'airlines': '运营航司', 'type': '机型', 
            
            'dep': '出发机场', 'arr': '到达机场', 'route': '航线', 
            'time_dep': '出发时刻', 'time_arr': '到达时刻', 'hour_dep': '出发时段', 
            'density_day': '日航班数', 'hour_comp': '时段竞争', 
            
            
            'count': '总计', 'route_count': '航线计数', 'type_count': '机型计数', 
            'date_flight_count': '日期计数', 'date_coll_count': '收集计数', 
            
            'price_total': '全价', 'price': '价格', 'price_rate': '折扣', 
            
            'ratio_daily': '当日系数', 
            'mid_price_rate': '折扣中位', 'avg_price_rate': '折扣平均', 
            'mid_ratio_price': '系数中位', 'avg_ratio_price': '系数平均', 
        }
        if isinstance(starting_date, date):
            starting_date = starting_date.toordinal()
        elif isinstance(starting_date, tuple) or isinstance(starting_date, list):
            starting_date = date(*starting_date).toordinal()
        self.__starting_date = starting_date if isinstance(starting_date, int) else 0
        self.__day_limit = day_limit if isinstance(day_limit, int) else 0
        self.__root = Path(root)
        
        self.__files: list[Path] = []
        self.__unlink: list[Path] = []
        
        self.__warn = 0
        
    '''EXCEL FORMAT'''
    #WORKBOOK           -> wb / file
    
    ##WORKSHEET         -> ws / sheet
    
    ###AAAAAAAAAA   ..A -> title
    ###BCCCCCCCCD   ..B -> header
    ###BCCCCCCCCD   ..C -> data
    ###BCCCCCCCCD   ..D -> tail
    ###BCCCCCCCCD   ..E -> footer
    ###EEEEEEEEEF   ..F -> (anything)
    
    '''DATA INPUT'''
    #FILE       -> merged_{root}.csv
    
    ##HEADER    -> date_flight   day_week airline type dep arr time_dep time_arr
    ###DATA     -> int (ordinal) str      str     str  str str isotime  isotime
    
    ##HEADER    -> price price_rate date_coll     day_adv hour_dep route
    ###DATA     -> int   float      int (ordinal) int     int      str
    
    def root(self, __root: Path | str = None, /) -> Path:
        '''Change root path if root path is given.
        
        Return seted root path in `Path`.'''
        if Path(__root).exists() and __root:
            self.__root = Path(__root)
        return self.__root
    
    
    def append_file(self, file: Path) -> Path:
        '''Append data to rebuilder for further processes.
        
        Return `None` for loading failure or none-excel file.'''
        try:
            datetime.fromisoformat(file.parent.name)
        except:
            print(f"WARN: {file.name} is not a standard path name of collecting date!")
            return None
        if file.match("*.xlsx"):
            self.__files.append(file)
            return file
        else:
            return None
    
    def append_folder(self, *paths: Path | str, count: int = 0) -> int:
        '''Load files from a folder, 
        whose name should be data's collecting date.
        
        Return the number of excels loaded.'''
        files = 0
        if len(paths) == 0:
            paths = []
            for path in self.__root.iterdir():
                if path.is_dir():
                    paths.append(path)
        for path in paths:
            path = Path(path)
            if self.__root != path.parent:
                path = self.__root / path
            try:
                if path.is_dir():
                    datetime.fromisoformat(path.name)
                else:
                    print(f"WARN: {path.name} should be an existing folder!")
                    self.__warn += 1
                    continue
            except:
                print(f"WARN: {path.name} is not a standard path name of collecting date!")
                self.__warn += 1
                continue
            
            for file in path.iterdir():
                if file.match("*.xlsx") and "_" not in file.name:
                    self.__files.append(file)
                    files += 1
                if count and files >= count:
                    break
        return files
    
    def append_zip(self, *paths: Path | str, file_name: str = "orig.zip") -> int:
        '''
        Append data from a zip file to process.
        
        - paths: `Path`, where to find and extract the zip file in `root`.
        
                default: all folders in the `root`
        
        - file: `Path` | `str`, the zip file path or name.
        
                default: `orig.zip` as a collection's extract.
        
        Return the number of excels loaded.
        '''
        files = 0
        if len(paths) == 0:
            paths = []
            for path in self.__root.iterdir():
                if path.is_dir():
                    paths.append(path)
        for path in paths:
            path = Path(path)
            if self.__root != path.parent:
                path = self.__root / path
            try:
                if path.is_dir():
                    datetime.fromisoformat(path.name)
                else:
                    print(f"WARN: {path.name} should be an existing folder!")
                    self.__warn += 1
                    continue
            except:
                print(f"WARN: {path.name} is not a standard path name of collecting date!")
                self.__warn += 1
                continue
            
            try:
                zip = ZipFile(path / Path(file_name), "r")
                zip.extractall(path)
                zip.close
            except:
                print(f"WARN: {file_name} cannot be loaded in", path.name)
                self.__warn += 1
                continue
            for item in path.iterdir():
                if item.match("*.xlsx") and "_" not in item.name:
                    files += 1
                    self.__files.append(item)
                    self.__unlink.append(item)
        return files
    
    
    def append_data(self, path: Path | str | None = None) -> DataFrame:
        '''Load / append merged pandas file
        
        Return appended data `DataFrame`'''
        if path == '' or path == None or path == Path():
            path = f'merged_{self.__root.name}.csv'
        print('loading data >>', Path(path).name)
        data = read_csv(Path(path))
        if not self.__header_min < set(data.keys()):
            print("ERROR: Required header missing!")
            return None
        elif not self.__header_req < set(data.keys()):
            print("ERROR: Merge data by merge method first!")
            return None
        if len(data) > 0:
            if self.__day_limit:
                data.drop(data[data['day_adv'] > self.__day_limit].index, inplace = True)
            if self.__starting_date:
                data.drop(data[data['date_flight'] < self.__starting_date].index, inplace = True)
            self.__merge = concat([data, self.__merge]) if len(self.__merge) else data
            return data
        else:
            print("ERROR: No valid data loaded!")
            return None
    
    
    def reset(self, unlink_file: bool = True, clear_rebuilt: bool = False) -> int:
        '''
        Clear all files in the data process queue
        -----
        - unlink_file: `True`, unlink excels zip file extracted
        - clear_rebuilt: `True`, clear all rebuilt data
        - Return current count of total warnings and reset to 0
        '''
        if unlink_file and len(self.__unlink):
            for file in self.__unlink:
                file.unlink()
        self.__files.clear()
        self.__unlink.clear()
        if clear_rebuilt:
            del self.__merge
            self.__merge = DataFrame()
            del self.__preprocess
            self.__preprocess = DataFrame()
        warn, self.__warn = self.__warn, 0 
        return warn
    
    @staticmethod
    def __diffrule(index: int | Literal['red', 'yellow', 'blue', 'green'] | tuple[str, str], 
                   type: str, op: str = None, formula = ()) -> Rule:
        '''index == 0 -> bold, even index -> red, odd index -> yellow, 
        str -> color, tuple[str, str] index -> color(font, patternfill)'''
        exp = {
            ">": "greaterThan", ">=": "greaterThanOrEqual", "<": "lessThan", 
            "<=": "lessThanOrEqual", "=": "equal", "==": "equal", "!=": "notEqual"}
        if isinstance(index, int):
            if index == 0:
                return Rule(type, operator = exp.get(op, op), stopIfTrue = False, formula = formula,
                    dxf = DifferentialStyle(font = Font(bold = "b")))
            elif index % 2:
                index = 'yellow'
            else:
                index = 'red'
        if isinstance(index, str):
            map = {'red': ("CC0000", "FFCCCC"), 'yellow': ("CC6600", "FFEBCD"), 
                   'blue': ("0000CC", "CCECFF"), 'green': ("008000", "CCFFCC")}
            index = map.get(index, ('000000', 'F0F0F0'))
        return Rule(type, operator = exp.get(op, op), stopIfTrue = False, formula = formula,
                    dxf = DifferentialStyle(font = Font(color = index[0]), \
                        fill = PatternFill(bgColor = index[1], fill_type = "solid")))
    
    def indexbook(self, data: DataFrame = None, airlines: bool = False):
        '''Return a workbook with route index aka menu'''
        data = self.__merge if data is None else data.get(['route', 'airline'])
        wb = Workbook()
        ws = wb.active
        ws.title = self.index_name
        ws.append(('航空公司' if airlines else '', '出发', '到达'))
        for idx in range(1, 4):
            ws.cell(1, idx).font = self.set_bold
            ws.cell(1, idx).alignment = self.set_align
        ws.auto_filter.ref, ws.freeze_panes = 'A1:B1', 'C2'
        
        cities, row = {}, 1
        for route in data['route'].unique():
            dep, arr = route.split('-', 1)
            if cities.get(dep):
                cities[dep].append(arr)
            else:
                cities[dep] = [arr]
        for city in cities.keys():
            cities[city].sort(key = lambda x: len(cities[x]), reverse = True)
        for dep in sorted(cities.keys(), key = lambda x: len(cities[x]), reverse = True):
            col = 2
            row += 1
            ws.cell(row, col, dep).alignment = self.set_align
            for arr in cities[dep]:
                col += 1
                cell = ws.cell(row, col, arr)
                cell.hyperlink = f"#'{dep}-{arr}'!A1"
                cell.font = self.set_hyperlink
                cell.alignment = self.set_align
        
        if airlines:
            ws.column_dimensions['A'].width = 12
            airlines = sorted(data['airline'].unique(), reverse = True, \
                key = lambda x: len(data.loc[data['airline'] == x]))
            for idx in range(len(airlines)):
                airline = airlines[idx]
                cell = ws.cell(idx + 2, 1, airline)
                cell.hyperlink = f"#'{airline}'!A1"
                cell.alignment = self.set_align
                cell.font = self.set_hyperlink
        return wb
    
    
    def merge(self) -> DataFrame:
        total = len(self.__files)
        if total == 0:
            raise ValueError("ERROR: NO FILE / DATA LOADED!")
        frame = []
        header = (
            'date_flight', 'day_week', 'airline', 'type', 'dep',                            #04
            'arr', 'time_dep', 'time_arr', 'price', 'price_rate')                           #09
        idct, percent = 0, -1
        for file in self.__files:
            idct += 1
            if percent != int(idct / total * 100):
                percent = int(idct / total * 100)
                print(f"\rmerging >> {percent:03d}", end = '%')
            date_coll = date.fromisoformat(file.parent.name).toordinal()
            data = read_excel(file, names = header).assign(date_coll = date_coll)    #10
            
            data['date_flight'] = data['date_flight'].map(lambda x: x.toordinal())
            if self.__starting_date:
                data.drop(data[data['date_flight'] < self.__starting_date].index, inplace = True)
            data['day_adv'] = data['date_flight'] - date_coll                               #11
            if self.__day_limit:
                data.drop(data[data['day_adv'] > self.__day_limit].index, inplace = True)
            data['hour_dep'] = data['time_dep'].map(lambda x: x.hour if x.hour else 24)     #12
            data['route'] = data['dep'].map(Airport) + data['arr'].map(Airport)             #13
            frame.append(data)
        print()
        return concat(frame)
    
    
    def dates(self, path: Path | str = Path(), file: str = '') -> None:
        '''Date overview by date of collect and date of flight
        
        Output separated excel with conditional formats'''
        
        if not len(self.__merge):
            self.__merge = self.merge()
        data = self.__merge.sort_values('date_flight')
        total_flights = sorted(date.fromordinal(ordinal) \
            for ordinal in data['date_flight'].unique())
        total_colls = sorted(date.fromordinal(ordinal) \
            for ordinal in data['date_coll'].unique())
        total, idct, percent = len(data), -1, 0
        
        if file == '' or file is None:
            file_coll = f"overview_{self.__root.name}_coll_dates.xlsx"
            file_flight = f"overview_{self.__root.name}_flight_dates.xlsx"
        else:
            file = file.replace(".xlsx", "")
            for char in file:
                if char >= u"一" and char <= u"龥":
                    file_coll = file + "_收集日期.xlsx"
                    file_flight = file + "_航班日期.xlsx"
                    break
            else:
                file_coll = file + "_coll_dates.xlsx"
                file_flight = file + "_flight_dates.xlsx"
        tstring = datetime.today().strftime('%H%M%S')
        if (Path(path) / Path(file_coll)).exists():
            file_coll = file_coll.replace(".xlsx", f"_{tstring}.xlsx")
        if (Path(path) / Path(file_flight)).exists():
            file_flight = file_flight.replace(".xlsx", f"_{tstring}.xlsx")
        
        '''Add index aka index'''
        wb_coll = Workbook()
        index_coll = wb_coll.active
        index_coll.title = self.index_name
        index_coll.column_dimensions['A'].width = 11
        index_coll.freeze_panes = 'E2'
        index_coll.append(["收集日期", "航班总数", "航线总数", "航班日数"] + total_flights)
        for idx in range(1, 5):
            index_coll.cell(1, idx).font = self.set_bold
            index_coll.cell(1, idx).alignment = self.set_align
        for idx in range(5, index_coll.max_column + 1):
            cell = index_coll.cell(1, idx)
            cell.number_format = self.set_date
            cell.alignment = self.set_align
            cell.hyperlink = f"{file_flight}#'{total_flights[idx - 5].strftime('%m-%d')}'!A2"
            cell.font = self.set_hyperlink
            cell.alignment = self.set_align
        
        wb_flight = Workbook()
        index_flight = wb_flight.active
        index_flight.title = self.index_name
        index_flight.column_dimensions['A'].width = 11
        index_flight.freeze_panes = 'E2'
        index_flight.append(["航班日期", "航班总数", "航线总数", "收集日数"] + total_colls)
        for idx in range(1, 5):
            index_flight.cell(1, idx).font = self.set_bold
            index_flight.cell(1, idx).alignment = self.set_align
        for idx in range(5, index_flight.max_column + 1):
            cell = index_flight.cell(1, idx)
            cell.number_format = self.set_date
            cell.alignment = self.set_align
            cell.hyperlink = f"{file_coll}#'{total_colls[idx - 5].strftime('%m-%d')}'!C3"
            cell.font = self.set_hyperlink
            cell.alignment = self.set_align
        
        
        '''Append date of collect data'''
        sheets_coll = data.groupby(["date_coll"])
        for coll_date, group in sheets_coll:
            sheet = date.fromordinal(coll_date)
            ws = wb_coll.create_sheet(sheet.strftime("%m-%d"))
            row = {}
            footers = ['平均', group["price_rate"].median(), group["price_rate"].mean()]
            header = [sheet, len(group), len(group["route"].unique()), len(group["date_flight"].unique())]
            flight_dates = sorted(group["date_flight"].unique())
            routes, rows = group.groupby(["route"]), group.groupby(["date_flight"])
            title = {
                'date': ["航线 \ 日期", "折扣中位", "折扣均值"], 
                'week': ["(星期)", None, None], 
                'adv': ["(提前天数)", None, None]}
            for ordinal in flight_dates:
                title['date'].append(date.fromordinal(ordinal))
                title['week'].append(date.fromordinal(ordinal).isoweekday())
                title['adv'].append(ordinal - coll_date)
                try:
                    footers.append(rows.get_group(ordinal)["price_rate"].mean())
                except:
                    footers.append(None)
            
            for route, group in routes:
                idct += len(group)
                if percent != int(idct / total * 25):
                    percent = int(idct / total * 25)
                    print(f"\rmerging dates >> {percent:03d}", end = '%')
                group.sort_values('date_flight', inplace = True)
                rows = group.groupby(["date_flight"])
                row[route] = [route, group["price_rate"].median(), group["price_rate"].mean()]
                for ordinal in flight_dates:
                    try:
                        row[route].append(rows.get_group(ordinal)["price_rate"].mean())
                    except:
                        row[route].append(None)
                del rows
            del routes
            
            '''Format sheet'''
            for item in title.values():
                ws.append(item)
            ws.cell(1, 1).hyperlink = f"#'{self.index_name}'!A1"
            ws.cell(1, 1).font = self.set_hyperlink
            for idx in range(4, ws.max_column + 1):
                ws.cell(1, idx).number_format = self.set_date
            
            for route in row.values():
                ws.append(route)
            for cols in ws.iter_cols(2, ws.max_column, 4, ws.max_row):
                for cell in cols:
                    cell.number_format = self.set_percent
            ws.freeze_panes = 'D4'
            ws.column_dimensions['A'].width = 14
            
            ws.append(footers)
            for idx in range(2, ws.max_column + 1):
                ws.cell(ws.max_row, idx).number_format = self.set_percent
            
            ratios = list(None for _ in total_flights)
            for item in title["date"][3:]:
                idx = total_flights.index(item)
                _idx = title["date"].index(item)
                ratios[idx] = f"='{ws.title}'!{ws.cell(ws.max_row, _idx + 1).coordinate}"
            index_coll.append(header + ratios)
            cell = index_coll.cell(index_coll.max_row, 1)
            cell.hyperlink = f"#'{ws.title}'!C3"
            cell.font = self.set_hyperlink
        del sheets_coll
        
        
        '''Append date of flight data'''
        sheets_flight = data.groupby(["date_flight"])
        for flight_date, group in sheets_flight:
            sheet = date.fromordinal(flight_date)
            ws = wb_flight.create_sheet(sheet.strftime("%m-%d"))
            row = {}
            footers = ['平均', group["price_rate"].median(), group["price_rate"].mean()]
            header = [sheet, len(group), len(group["route"].unique()), len(group["date_coll"].unique())]
            coll_dates = sorted(group["date_coll"].unique())
            routes, rows = group.groupby(["route"]), group.groupby(["date_coll"])
            title = ["航线 \ 收集", "折扣中位", "折扣均值"]
            for ordinal in coll_dates:
                title.append(date.fromordinal(ordinal))
                try:
                    footers.append(rows.get_group(ordinal)["price_rate"].mean())
                except:
                    footers.append(None)
            
            for route, group in routes:
                idct += len(group)
                if percent != int(idct / total * 25 + 25):
                    percent = int(idct / total * 25 + 25)
                    print(f"\rmerging dates >> {percent:03d}", end = '%')
                group.sort_values('date_coll', inplace = True)
                rows = group.groupby(["date_coll"])
                row[route] = [route, group["price_rate"].median(), group["price_rate"].mean()]
                for ordinal in coll_dates:
                    try:
                        row[route].append(rows.get_group(ordinal)["price_rate"].mean())
                    except:
                        row[route].append(None)
                del rows
            del routes
            
            '''Format sheet'''
            ws.append(title)
            ws.cell(1, 1).hyperlink = f"#'{self.index_name}'!A1"
            ws.cell(1, 1).font = self.set_hyperlink
            for idx in range(2, ws.max_column + 1):
                ws.cell(1, idx).number_format = self.set_date
            
            for route in row.values():
                ws.append(route)
            for cols in ws.iter_cols(2, ws.max_column, 2, ws.max_row):
                for cell in cols:
                    cell.number_format = self.set_percent
            ws.freeze_panes = 'D2'
            ws.column_dimensions['A'].width = 14
            
            ws.append(footers)
            for idx in range(2, ws.max_column + 1):
                ws.cell(ws.max_row, idx).number_format = self.set_percent
            
            ratios = list(None for _ in total_colls)
            for item in title[3:]:
                idx = total_colls.index(item)
                _idx = title.index(item)
                ratios[idx] = f"='{ws.title}'!{ws.cell(ws.max_row, _idx + 1).coordinate}"
            index_flight.append(header + ratios)
            cell = index_flight.cell(index_flight.max_row, 1)
            cell.hyperlink = f"#'{ws.title}'!A2"
            cell.font = self.set_hyperlink
        del sheets_flight
        
        
        '''Sheets condition format'''
        for cols in index_coll.iter_cols(5, index_coll.max_column, 2, index_coll.max_row):
            for cell in cols:
                cell.number_format = self.set_percent
        total, idct = len(wb_coll.sheetnames), 0
        for ws in wb_coll:
            idct += 1
            if percent != int(idct / total * 25 + 50):
                percent = int(idct / total * 25 + 50)
                print(f"\rmerging dates >> {percent:03d}", end = '%')
            if ws.title == self.index_name:
                continue
            cell = ws.cell(3, 3, "返回索引")
            cell.hyperlink = f"#'{self.index_name}'!A1"
            cell.font = self.set_hyperlink
            med_string = f"B4:B{ws.max_row - 1}"
            med_rule = self.__diffrule(0, 'cellIs', '>', [f"$B${ws.max_row}"])
            ws.conditional_formatting.add(med_string, med_rule)
            avg_string = f"C4:C{ws.max_row - 1}"
            avg_rule = self.__diffrule(0, 'cellIs', '>', [f"$C${ws.max_row}"])
            ws.conditional_formatting.add(avg_string, avg_rule)
            for row in range(4, ws.max_row):
                rstring = f"{ws.cell(row, 4).coordinate}:{ws.cell(row, ws.max_column).coordinate}"
                ws.conditional_formatting.add(rstring, self.__diffrule(row, 'cellIs', '>', [f"$C${row}"]))
        
        
        for cols in index_flight.iter_cols(5, index_flight.max_column, 2, index_flight.max_row):
            for cell in cols:
                cell.number_format = self.set_percent
        total, idct = len(wb_flight.sheetnames), 0
        for ws in wb_flight:
            idct += 1
            if percent != int(idct / total * 25 + 75):
                percent = int(idct / total * 25 + 75)
                print(f"\rmerging dates >> {percent:03d}", end = '%')
            if ws.title == self.index_name:
                continue
            med_string = f"B4:B{ws.max_row - 1}"
            med_rule = self.__diffrule(0, 'cellIs', '>', [f"$B${ws.max_row}"])
            ws.conditional_formatting.add(med_string, med_rule)
            avg_string = f"C4:C{ws.max_row - 1}"
            avg_rule = self.__diffrule(0, 'cellIs', '>', [f"$C${ws.max_row}"])
            ws.conditional_formatting.add(avg_string, avg_rule)
            for row in range(2, ws.max_row):
                rstring = f"{ws.cell(row, 4).coordinate}:{ws.cell(row, ws.max_column).coordinate}"
                ws.conditional_formatting.add(rstring, self.__diffrule(row, 'cellIs', '>', [f"$C${row}"]))
        
        '''Output merged data'''
        print("\r saving")
        wb_coll.save(Path(path) / Path(file_coll))
        wb_coll.close
        wb_flight.save(Path(path) / Path(file_flight))
        wb_flight.close
    
    
    def routes(self, path: Path | str = Path(), file: str = ''):
        '''Route overview'''
        if not len(self.__merge):
            self.__merge = self.merge()
        data = self.__merge.copy()
        
        if 'density_day' not in data.keys():
            data['density_day'] = data.groupby(["date_flight", \
                "route", "date_coll"])['date_flight'].transform("count")
        if 'ratio_daily' not in data.keys():
            data['ratio_daily'] = data["price_rate"] / data.groupby(["date_coll", \
                "date_flight", "route"])["price_rate"].transform("mean")
        if 'hour_comp' not in data.keys():
            data['hour_comp'] = data.groupby(["date_coll","date_flight", \
                "hour_dep", "route"])["airline"].transform("nunique")
        
        overview, total = data.groupby(["route"]), len(data)
        title = [
            '航线', '总计', '航班日数', '收集日数', '机型数量', '全价', 
            '折扣平均', '折扣中位', '运营航司', '时段竞争', '日航班数']
        title_adv = {"overview": title + sorted(data['day_adv'].unique())}
        title_hour = title + list(range(5, 25))
        title_date = title + list(date.fromordinal(ordinal) \
            for ordinal in data['date_flight'].unique())
        title_coll = {}
        
        route_density = {}
        route_ratio = {}
        route_comp = {}
        route_adv_mean = {}
        route_adv_std = {}
        route_date = {}
        route_date_mean = {}
        route_date_std = {}
        route_coll = {}
        headers = {}
        footers = {}
        idct, percent = 0, -1
        for name, group in overview:
            idct += len(group)
            if percent != int(idct / total * 80):
                percent = int(idct / total * 80)
                print(f"\rmerging routes >> {percent:03d}", end = '%')
            
            headers[name] = [
                name, len(group), group['date_flight'].nunique(), 
                group['date_coll'].nunique(), group['type'].nunique(), 
                Route.fromformat(name).airfare, group['price_rate'].mean(), 
                group['price_rate'].median(), group['airline'].nunique(), 
                round(group['hour_comp'].mean(), 2), round(group['density_day'].mean())]
            footers[name] = {'date': {}, 'day': [], 'coll': []}
            route_comp[name] = []
            route_density[name] = []
            route_ratio[name] = []
            route_adv_mean[name] = []
            route_adv_std[name] = []
            route_date[name] = {}
            route_date_mean[name] = []
            route_date_std[name] = []
            route_coll[name] = {}
            title_adv[name] = sorted(group['day_adv'].unique())
            title_coll[name] = sorted(group['date_coll'].unique(), reverse = True)
            
            for hour in range(5, 25):
                hour_airline = group.loc[group['hour_dep'] == hour].get('hour_comp', 0).mean()
                route_comp[name].append(hour_airline if hour_airline else None)
                hour_count = group['hour_dep'].value_counts().get(hour, 0)
                route_density[name].append(round(hour_count / len(group['hour_dep']), 2) \
                    if hour_count else None)
                hour_ratio = group.loc[group['hour_dep'] == hour].get('ratio_daily', 0).mean()
                route_ratio[name].append(round(hour_ratio, 2) if hour_ratio else None)
            for day in data['day_adv'].unique():
                mean = group.loc[group['day_adv'] == day].get('price_rate', 0).mean()
                if mean:
                    route_adv_mean[name].append(mean)
                else:
                    route_adv_mean[name].append(None)
                    continue
                std = group.loc[group['day_adv'] == day].get('price_rate', 0).std()
                route_adv_std[name].append(std if std else None)
            for ordinal in data['date_flight'].unique():
                mean = group.loc[group['date_flight'] == ordinal].get('price_rate', 0).mean()
                if mean:
                    route_date_mean[name].append(mean)
                    footers[name]['date'][ordinal] = [mean, 
                        group.loc[group['date_flight'] == ordinal].get('price_rate', 0).median()]
                else:
                    route_date_mean[name].append(None)
                    continue
                std = group.loc[group['date_flight'] == ordinal].get('price_rate', 0).std()
                route_date_std[name].append(std if std else None)
            for day in group['date_flight'].unique():
                route_date[name][day] = list(None for _ in title_adv[name])
                route_coll[name][day] = list(None for _ in title_coll[name])
            for (ordinal, day), _group in group.groupby(['date_flight', 'day_adv']):
                route_date[name][ordinal][title_adv[name].index(day)] = _group['price_rate'].mean()
            for (ordinal, day), _group in group.groupby(['date_flight', 'date_coll']):
                route_coll[name][ordinal][title_coll[name].index(day)] = _group['price_rate'].mean()
            for day, _group in group.groupby(['day_adv']):
                footers[name]['day'].append(_group['price_rate'].mean())
            for day, _group in group.groupby(['date_coll']):
                footers[name]['coll'].append(_group['price_rate'].mean())
            footers[name]['coll'].reverse()
        
        wb = self.indexbook()
        for sheet in ('时刻密度', '时刻竞争', '时刻系数'):
            ws = wb.create_sheet(sheet)
            ws.append(title_hour)
            for idx in range(1, len(title_hour) + 1):
                ws.cell(1, idx).font = self.set_bold
                ws.cell(1, idx).alignment = self.set_align
        for sheet in ('单日平均折扣', '单日标准差'):
            ws = wb.create_sheet(sheet)
            ws.append(title_date)
            for idx in range(1, len(title_date) + 1):
                if idx > len(title):
                    ws.cell(1, idx).number_format = self.set_date
                ws.cell(1, idx).font = self.set_bold
                ws.cell(1, idx).alignment = self.set_align
        for sheet in ('提前平均折扣', '提前标准差'):
            ws = wb.create_sheet(sheet)
            ws.append(title_adv["overview"])
            for idx in range(1, len(title_adv["overview"]) + 1):
                ws.cell(1, idx).font = self.set_bold
                ws.cell(1, idx).alignment = self.set_align
        
        for name in headers.keys():
            wb['时刻密度'].append(headers[name] + route_density[name])
            wb['时刻竞争'].append(headers[name] + route_comp[name])
            wb['时刻系数'].append(headers[name] + route_ratio[name])
            wb['提前平均折扣'].append(headers[name] + route_adv_mean[name])
            wb['提前标准差'].append(headers[name] + route_adv_std[name])
            wb['单日平均折扣'].append(headers[name] + route_date_mean[name])
            wb['单日标准差'].append(headers[name] + route_date_std[name])
        for ws in wb:
            if ws.title == self.index_name:
                continue
            ws.freeze_panes = 'L2'
            ws.column_dimensions['A'].width = 14
            for idx in range(2, ws.max_row + 1):
                ws.cell(idx, 1).hyperlink = f"#'{ws.cell(idx, 1).value}'!B1"
                ws.cell(idx, 1).font = self.set_hyperlink
                ws.cell(idx, 7).number_format = self.set_percent
                ws.cell(idx, 8).number_format = self.set_percent
        
        total, idct = len(headers), 0
        for name in headers.keys():
            idct += 1
            if percent != int(idct / total * 20 + 80):
                percent = int(idct / total * 20 + 80)
                print(f"\rmerging routes >> {percent:03d}", end = '%')
            
            '''Append data by days advanced'''
            ws = wb.create_sheet(name)
            ws.append(['日期\提前天', '星期', '折扣平均', '折扣中位'] + title_adv[name])
            for idx in range(2, 5):
                cell = ws.cell(1, idx)
                cell.alignment = self.set_align
                cell.font = self.set_bold
            for ordinal in route_date[name]:
                day = date.fromordinal(ordinal)
                ws.append([day, day.isoweekday()] + \
                    footers[name]['date'][ordinal] + route_date[name][ordinal])
            ws.cell(1, 1).hyperlink = f"#'{self.index_name}'!A1"
            ws.cell(1, 1).font = self.set_hyperlink
            ws.freeze_panes = 'E2'
            ws.column_dimensions['A'].width = 12
            ws.column_dimensions["B"].width = 6
            ws.append(['按收集日排序', '返程', '返回索引', '平均'] + footers[name]['day'])
            for cols in ws.iter_cols(3, ws.max_column, 2, ws.max_row):
                for cell in cols:
                    cell.number_format = self.set_percent
            
            '''Sheets condition format and averages'''
            med_string = f"D2:D{ws.max_row - 1}"
            med_rule = self.__diffrule(0, 'aboveAverage')
            ws.conditional_formatting.add(med_string, med_rule)
            avg_string = f"C2:C{ws.max_row - 1}"
            avg_rule = self.__diffrule(0, 'aboveAverage')
            ws.conditional_formatting.add(avg_string, avg_rule)
            for idx in range(5, ws.max_column + 1):
                rstring = f"{ws.cell(2, idx).coordinate}:{ws.cell(ws.max_row - 1, idx).coordinate}"
                rules = (idx, 'cellIs', '>', [footers[name]['day'][idx - 5]])
                ws.conditional_formatting.add(rstring, self.__diffrule(*rules))
            
            cell, (dep, arr) = ws.cell(ws.max_row, 1), name.split('-', 1)
            cell.hyperlink = f"#'({dep}-{arr})'!A2"
            cell.font = self.set_hyperlink
            cell = ws.cell(ws.max_row, 2)
            cell.hyperlink = f"#'{arr}-{dep}'!A2"
            cell.font = self.set_hyperlink
            cell = ws.cell(ws.max_row, 3)
            cell.hyperlink = f"#'{self.index_name}'!A1"
            cell.font = self.set_hyperlink
            
            '''Append data by date of collect'''
            ws = wb.create_sheet(f"({name})")
            ws.append(['日期\收集日', '星期', '折扣平均', '折扣中位'] + \
                list(date.fromordinal(ordinal) for ordinal in title_coll[name]))
            for idx in range(2, 5):
                cell = ws.cell(1, idx)
                cell.alignment = self.set_align
                cell.font = self.set_bold
            for idx in range(5, ws.max_column + 1):
                cell = ws.cell(1, idx)
                cell.number_format = self.set_date
                cell.alignment = self.set_align
            for ordinal in route_coll[name]:
                day = date.fromordinal(ordinal)
                ws.append([day, day.isoweekday()] + \
                    footers[name]['date'][ordinal] + route_coll[name][ordinal])
            ws.cell(1, 1).hyperlink = f"#'{self.index_name}'!A1"
            ws.cell(1, 1).font = self.set_hyperlink
            ws.freeze_panes = 'E2'
            ws.column_dimensions['A'].width = 12
            ws.column_dimensions["B"].width = 6
            ws.append(['按提前天排序', '返程', '返回索引', '平均'] + footers[name]['coll'])
            for cols in ws.iter_cols(3, ws.max_column, 2, ws.max_row):
                for cell in cols:
                    cell.number_format = self.set_percent
            
            '''Sheets condition format and averages'''
            med_string = f"D2:D{ws.max_row - 1}"
            med_rule = self.__diffrule(0, 'aboveAverage')
            ws.conditional_formatting.add(med_string, med_rule)
            avg_string = f"C2:C{ws.max_row - 1}"
            avg_rule = self.__diffrule(0, 'aboveAverage')
            ws.conditional_formatting.add(avg_string, avg_rule)
            for idx in range(5, ws.max_column + 1):
                rstring = f"{ws.cell(2, idx).coordinate}:{ws.cell(ws.max_row - 1, idx).coordinate}"
                rules = (idx, 'cellIs', '>', [footers[name]['coll'][idx - 5]])
                ws.conditional_formatting.add(rstring, self.__diffrule(*rules))
            
            cell, (dep, arr) = ws.cell(ws.max_row, 1), name.split('-', 1)
            cell.hyperlink = f"#'{dep}-{arr}'!A2"
            cell.font = self.set_hyperlink
            cell = ws.cell(ws.max_row, 2)
            cell.hyperlink = f"#'({arr}-{dep})'!A2"
            cell.font = self.set_hyperlink
            cell = ws.cell(ws.max_row, 3)
            cell.hyperlink = f"#'{self.index_name}'!A1"
            cell.font = self.set_hyperlink
            
        '''Output merged data'''
        if file == '' or file is None:
            file = f"overview_{self.__root.name}_routes.xlsx"
        elif not file.endswith(".xlsx"):
            file += ".xlsx"
        if not isinstance(path, Path):
            path = Path(path)
        path.mkdir(parents = True, exist_ok = True)
        if (path / Path(file)).exists():
            time = datetime.today().strftime("%H%M%S")
            file = file.replace(".xlsx", f"_{time}.xlsx")
        print("\r saving")
        wb.save(path / Path(file))
        wb.close()
    
    
    def airlines(self, path: Path | str = Path(), file: str = ''):
        '''Airline overview'''
        if not len(self.__merge):
            self.__merge = self.merge()
        data = self.__merge.copy()
        
        if 'ratio_daily' not in data.keys():
            data['ratio_daily'] = data["price_rate"] / data.groupby(["date_coll", \
                "date_flight", "route"])["price_rate"].transform("mean")
        
        overview, total = data.groupby(["airline"]), len(data)
        title = [
            '航空公司', '总计', '航班日数', '收集日数', 
            '航线数量', '机型数量', '系数平均', '系数中位']
        title_hour = title + list(range(5, 25))
        title_route = title + sorted(data['route'].unique())
        title_dep = title + sorted(data['dep'].unique())
        headers = {}
        routes = {}
        airlines = {}
        hour_density = {}
        hour_ratio = {}
        route_count = {}
        route_density = {}
        route_ratio = {}
        dep_ap = {}
        
        idct, percent = 0, -1
        for name, group in overview:
            headers[name] = [
                name, len(group), group['date_flight'].nunique(), group['date_coll'].nunique(), 
                group['route'].nunique(), group['type'].nunique(), group['ratio_daily'].mean(), 
                group['ratio_daily'].median()]
            route_count[name] = sorted(group['route'].unique())
            hour_density[name] = []
            hour_ratio[name] = []
            airlines[name] = {}
            route_density[name] = list(None for _ in data['route'].unique())
            route_ratio[name] = list(None for _ in data['route'].unique())
            dep_ap[name] = list(None for _ in data['dep'].unique())
            
            idct += len(group) / 4
            if percent != int(idct / total * 100):
                percent = int(idct / total * 100)
                print(f"\rmerging airlines >> {percent:03d}", end = '%')
            for hour in range(5, 25):
                count = group['hour_dep'].value_counts().get(hour, 0)
                hour_density[name].append(round(count / len(group['hour_dep']), 2) \
                    if count else None)
                ratio = group.loc[group['hour_dep'] == hour].get('ratio_daily').mean()
                hour_ratio[name].append(ratio if ratio else None)
            
            idct += len(group) / 4
            if percent != int(idct / total * 100):
                percent = int(idct / total * 100)
                print(f"\rmerging airlines >> {percent:03d}", end = '%')
            for route in group['route'].unique():
                airlines[name][route] = []
                idx = title_route.index(route) - len(title)
                count = group['route'].value_counts().get(route)
                days = len(group.loc[group['route'] == route]\
                    [['date_coll', 'date_flight']].drop_duplicates())
                route_density[name][idx] = count / days
                ratio = group.loc[group['route'] == route].get('ratio_daily').mean()
                route_ratio[name][idx] = ratio
            
            idct += len(group) / 4
            if percent != int(idct / total * 100):
                percent = int(idct / total * 100)
                print(f"\rmerging airlines >> {percent:03d}", end = '%')
            for dep in group['dep'].unique():
                idx = title_dep.index(dep) - len(title)
                count = group['dep'].value_counts().get(dep) 
                dep_ap[name][idx] = count / len(group.loc[group['dep'] == dep]\
                    [['date_coll', 'date_flight']].drop_duplicates())
            
            idct += len(group) / 4
            if percent != int(idct / total * 100):
                percent = int(idct / total * 100)
                print(f"\rmerging airlines >> {percent:03d}", end = '%')
            for route, _group in group.groupby(['route']):
                airlines[name][route] = []
                for hour in range(5, 25):
                    ratio = _group.loc[_group['hour_dep'] == hour].get('ratio_daily').mean()
                    airlines[name][route].append(ratio if ratio else None)
            for route in group['route'].unique():
                if routes.get(route):
                    routes[route][name] = airlines[name][route].copy()
                else:
                    routes[route] = {name: airlines[name][route].copy()}
        
        wb = self.indexbook(airlines = True)
        for sheet in ('航线密度', '航线系数'):
            ws = wb.create_sheet(sheet)
            ws.append(title_route)
            for idx in range(1, len(title) + 1):
                ws.cell(1, idx).font = self.set_bold
                ws.cell(1, idx).alignment = self.set_align
            for idx in range(len(title) + 1, len(title_route) + 1):
                cell = ws.cell(1, idx)
                cell.hyperlink = f"#'{cell.value}'!A1"
                cell.font = self.set_hyperlink
                cell.alignment = self.set_align
                ws.column_dimensions[cell.coordinate[:1]].width = 10
        
        for sheet in ('时刻密度', '时刻系数'):
            ws = wb.create_sheet(sheet)
            ws.append(title_hour)
            for idx in range(1, len(title_hour) + 1):
                ws.cell(1, idx).font = self.set_bold
                ws.cell(1, idx).alignment = self.set_align
        
        ws = wb.create_sheet('机场计数')
        ws.append(title_dep)
        for idx in range(1, len(title_dep) + 1):
            ws.cell(1, idx).font = self.set_bold
            ws.cell(1, idx).alignment = self.set_align
        
        for name in headers.keys():
            wb['航线密度'].append(headers[name] + route_density[name])
            wb['航线系数'].append(headers[name] + route_ratio[name])
            wb['时刻密度'].append(headers[name] + hour_density[name])
            wb['时刻系数'].append(headers[name] + hour_ratio[name])
            wb['机场计数'].append(headers[name] + dep_ap[name])
        for sheet in ('时刻系数', '航线系数'):
            ws = wb[sheet]
            for cols in ws.iter_cols(len(title) + 1, ws.max_column, 2, ws.max_row):
                for cell in cols:
                    cell.number_format = self.set_percent
        
        for ws in wb:
            if ws.title == self.index_name:
                continue
            ws.freeze_panes = 'I2'
            ws.column_dimensions['A'].width = 13
            for idx in range(2, ws.max_row + 1):
                ws.cell(idx, 1).hyperlink = f"#'{ws.cell(idx, 1).value}'!A1"
                ws.cell(idx, 1).font = self.set_hyperlink
                ws.cell(idx, len(title)).number_format = self.set_percent
                ws.cell(idx, len(title) - 1).number_format = self.set_percent
        
        '''Airline details'''
        for airline in airlines.keys():
            ws = wb.create_sheet(airline)
            ws.append(['航线'] + list(range(5, 25)))
            ws.freeze_panes = 'B2'
            ws.column_dimensions['A'].width = 14
            for route in airlines[airline].keys():
                ws.append([route] + airlines[airline][route])
            for idx in range(2, ws.max_row + 1):
                ws.cell(idx, 1).font = self.set_hyperlink
                ws.cell(idx, 1).hyperlink = f"#'{ws.cell(idx, 1).value}'!A1"
            for idx in range(2, 22):
                ws.cell(1, idx).alignment = self.set_align
                ws.cell(1, idx).font = self.set_bold
            for cols in ws.iter_cols(2, ws.max_column, 2, ws.max_row):
                for cell in cols:
                    cell.number_format = self.set_percent
            cell = ws.cell(ws.max_row + 1, 1, '返回索引')
            cell.hyperlink = f"#'{self.index_name}'!A1"
            cell.font = self.set_hyperlink
            ws.cell(1, 1).hyperlink = f"#'{self.index_name}'!A1"
            ws.cell(1, 1).font = Font(u = "single", color = "0070C0", bold = "b")
            ws.cell(1, 1).alignment = self.set_align
        
        '''Route details'''
        for route in routes.keys():
            ws = wb.create_sheet(route)
            ws.append(['航司'] + list(range(5, 25)))
            ws.freeze_panes = 'B2'
            ws.column_dimensions['A'].width = 13
            for airline in routes[route].keys():
                ws.append([airline] + routes[route][airline])
            for idx in range(2, ws.max_row + 1):
                ws.cell(idx, 1).font = self.set_hyperlink
                ws.cell(idx, 1).hyperlink = f"#'{ws.cell(idx, 1).value}'!A1"
            for idx in range(2, 22):
                ws.cell(1, idx).alignment = self.set_align
                ws.cell(1, idx).font = self.set_bold
            for cols in ws.iter_cols(2, ws.max_column, 2, ws.max_row):
                for cell in cols:
                    cell.number_format = self.set_percent
            cell = ws.cell(ws.max_row + 1, 1, '返回索引')
            cell.hyperlink = f"#'{self.index_name}'!A1"
            cell.font = self.set_hyperlink
            cell = ws.cell(ws.max_row + 1, 1, '返程')
            route = route.split('-', 1)
            cell.hyperlink = f"#'{route[1]}-{route[0]}'!A2"
            cell.font = self.set_hyperlink
            ws.cell(1, 1).hyperlink = f"#'{self.index_name}'!A1"
            ws.cell(1, 1).font = Font(u = "single", color = "0070C0", bold = "b")
            ws.cell(1, 1).alignment = self.set_align
        
        '''Output merged data'''
        if file == '' or file is None:
            file = f"overview_{self.__root.name}_airlines.xlsx"
        elif not file.endswith(".xlsx"):
            file += ".xlsx"
        if not isinstance(path, Path):
            path = Path(path)
        path.mkdir(parents = True, exist_ok = True)
        if (path / Path(file)).exists():
            time = datetime.today().strftime("%H%M%S")
            file = file.replace(".xlsx", f"_{time}.xlsx")
        print("\r saving")
        wb.save(path / Path(file))
        wb.close()

    def month(self, year: int = 0, month: int = 0, limit: int = 5, 
              path: Path | str = Path(), file: str = ''):
        '''Calculate pearson correlation coefficient of tickets rate before `year`.`month`'''
        filterwarnings("ignore")
        if not len(self.__merge):
            self.__merge = self.merge()
        max_coll = date.fromordinal(self.__merge['date_coll'].max())
        month = month if month else max_coll.month
        year = year if year else max_coll.year
        max_coll = (datetime(year, month, 1) - timedelta(1)).date()
        data = self.__merge.drop(
            self.__merge[self.__merge['date_coll'] >= max_coll.toordinal()].index).reset_index()
        data.drop(data[data['day_adv'] <= limit].index, inplace = True)
        
        wb = self.indexbook(airlines = True)
        rules = (
            self.__diffrule('red', 'cellIs', '>', [0.9]),
            self.__diffrule('yellow', 'cellIs', '>', [0.8]),
            self.__diffrule('', 'cellIs', '>', [0.5]),
            self.__diffrule('blue', 'cellIs', '<', [-0.9]), 
            self.__diffrule('green', 'cellIs', '<', [-0.8]), 
            self.__diffrule('', 'cellIs', '<', [-0.5]))
        percent, idct = -1, 0
        total = len(data) * 2
        data['date_flight'] = data['date_flight'].map(date.fromordinal)
        data['date_coll'] = data['date_coll'].map(date.fromordinal)
        data['target'] = data['date_flight'].map(lambda x: 1 if x.month == month else 0)
        
        overview = wb.create_sheet('总览')
        ov_title = sorted(data['airline'].unique(), reverse = True, 
            key = lambda x: len(self.__merge.loc[self.__merge['airline'] == x]))
        ov_index = sorted(data['route'].unique(), reverse = True, 
            key = lambda x: len(self.__merge.loc[self.__merge['route'] == x]))
        overview.cell(1, 1, '航线').font = self.set_bold
        overview.column_dimensions['A'].width = 13
        overview.freeze_panes = 'B2'
        for idx in range(len(ov_title)):
            cell = overview.cell(1, idx + 2, ov_title[idx])
            cell.font = self.set_hyperlink
            cell.hyperlink = f"#'{ov_title[idx]}'!A1"
        for idx in range(len(ov_index)):
            cell = overview.cell(idx + 2, 1, ov_index[idx])
            cell.font = self.set_hyperlink
            cell.hyperlink = f"#'{ov_index[idx]}'!A1"
        
        for airline, airlines in data.groupby(['airline']):
            col = 2 + ov_title.index(airline)
            ws = wb.create_sheet(airline)
            routes = sorted(airlines['route'].unique(), reverse = True, 
                key = lambda x: len(airlines.loc[airlines['route'] == x]))
            ws.append(['采集日期', '相关平均', '强正相关', '正相关', '负相关', '强负相关'] + routes)
            ws.auto_filter.ref = 'A1:F1'
            for idx in range(ws.max_column):
                cell = ws.cell(1, idx + 1)
                cell.alignment = self.set_align
                if idx >= 6:
                    cell.hyperlink = f"#'{cell.value}'!A1"
                    cell.font = self.set_hyperlink
                else:
                    cell.font = self.set_bold
            ws.freeze_panes = 'G2'
            for coll, colls in airlines.groupby(['date_coll']):
                idct += len(colls)
                if percent != int(idct / total * 100):
                    percent = int(idct / total * 100)
                    print(f"\rmonth corr >> {percent:03d}", end = '%')
                if len(colls) <= 1 or coll >= max_coll - timedelta(limit):
                    continue
                corr = colls['price_rate'].corr(colls['target'])
                output = [coll, corr, 0, 0, 0, 0] + list(None for _ in range(len(routes)))
                for route, group in colls.groupby(['route']):
                    corr = group['price_rate'].corr(group['target'], 'pearson')
                    output[routes.index(route) + 6] = round(corr, 4)
                    if corr > 0.8:
                        output[2] += 1
                    elif corr > 0:
                        output[3] += 1
                    elif corr < -0.8:
                        output[5] += 1
                    elif corr < 0:
                        output[4] += 1
                ws.append(output)
                ws.cell(ws.max_row, 1).number_format = self.set_date
            
            rstring = f"G2:{ws.cell(ws.max_row, ws.max_column).coordinate}"
            for rule in rules:
                ws.conditional_formatting.add(rstring, rule)
            ws.conditional_formatting.add(
                f"C2:C{ws.max_row}", self.__diffrule(0, 'cellIs', '>=', [1]))
            ws.conditional_formatting.add(
                f"F2:F{ws.max_row}", self.__diffrule(0, 'cellIs', '>=', [1]))
            
            cell = ws.cell(1, 1)
            cell.hyperlink = f"#'{self.index_name}'!A1"
            cell.font = self.set_hyperlink
            cell = ws.cell(ws.max_row + 1, 1, '返回索引')
            cell.hyperlink = f"#'{self.index_name}'!A1"
            cell.font = self.set_hyperlink
            for idx in range(7, ws.max_column + 1):
                cell = ws.cell(ws.max_row, idx)
                coord = cell.coordinate
                _col = coord.strip(str(ws.max_row))
                cell.value = f'=AVERAGE({_col}2:{_col}{ws.max_row - 1})'
                overview.cell(2 + ov_index.index(ws.cell(1, idx).value), col).value = \
                    f'={airline}!{coord}'
        
        rstring = f"B2:{overview.cell(overview.max_row, overview.max_column).coordinate}"
        for rule in rules:
            overview.conditional_formatting.add(rstring, rule)
        
        for route, routes in data.groupby(['route']):
            idct += len(routes)
            if percent != int(idct / total * 100):
                percent = int(idct / total * 100)
                print(f"\rmonth corr >> {percent:03d}", end = '%')
            ws = wb.create_sheet(route)
            airlines = sorted(routes['airline'].unique(), reverse = True, 
                key = lambda x: len(routes.loc[routes['airline'] == x]))
            ws.append(['采集日期', '相关平均', '强正相关', '正相关', '负相关', '强负相关'] + airlines)
            ws.auto_filter.ref = 'A1:F1'
            for idx in range(ws.max_column):
                cell = ws.cell(1, idx + 1)
                cell.alignment = self.set_align
                if idx >= 6:
                    cell.hyperlink = f"#'{cell.value}'!A1"
                    cell.font = self.set_hyperlink
                else:
                    cell.font = self.set_bold
            ws.freeze_panes = 'G2'
            for coll, colls in routes.groupby(['date_coll']):
                if len(colls) <= 1 or coll >= max_coll - timedelta(limit):
                    continue
                corr = colls['price_rate'].corr(colls['target'])
                output = [coll, corr, 0, 0, 0, 0] + list(None for _ in range(len(airlines)))
                for airline, group in colls.groupby(['airline']):
                    corr = group['price_rate'].corr(group['target'], 'pearson')
                    output[airlines.index(airline) + 6] = round(corr, 4)
                    if corr > 0.8:
                        output[2] += 1
                    elif corr > 0:
                        output[3] += 1
                    elif corr < -0.8:
                        output[5] += 1
                    elif corr < 0:
                        output[4] += 1
                ws.append(output)
                ws.cell(ws.max_row, 1).number_format = self.set_date
            rstring = f"G2:{ws.cell(ws.max_row, ws.max_column).coordinate}"
            for rule in rules:
                ws.conditional_formatting.add(rstring, rule)
            ws.conditional_formatting.add(
                f"C2:C{ws.max_row}", self.__diffrule(0, 'cellIs', '>=', [1]))
            ws.conditional_formatting.add(
                f"F2:F{ws.max_row}", self.__diffrule(0, 'cellIs', '>=', [1]))
            
            cell = ws.cell(1, 1)
            cell.hyperlink = f"#'{self.index_name}'!A1"
            cell.font = self.set_hyperlink
            cell = ws.cell(ws.max_row + 1, 1, '返回索引')
            cell.hyperlink = f"#'{self.index_name}'!A1"
            cell.font = self.set_hyperlink
            cell = ws.cell(ws.max_row, 2, '返程')
            route = route.split('-', 1)
            cell.hyperlink = f"#'{route[1]}-{route[0]}'!A1"
            cell.font = self.set_hyperlink
            for idx in range(7, ws.max_column + 1):
                cell = ws.cell(ws.max_row, idx)
                coord = cell.coordinate
                col = coord.strip(str(ws.max_row))
                cell.value = f'=AVERAGE({col}2:{col}{ws.max_row - 1})'
            
        print("\rsaving    ")
        filterwarnings("default")
        if file == '' or file is None:
            file = f"corr_month_{self.__root.name}.xlsx"
        elif not file.endswith(".xlsx"):
            file += ".xlsx"
        if not isinstance(path, Path):
            path = Path(path)
        path.mkdir(parents = True, exist_ok = True)
        if (path / Path(file)).exists():
            time = datetime.today().strftime("%H%M%S")
            file = file.replace(".xlsx", f"_{time}.xlsx")
        wb.save(path / Path(file))
        wb.close()
    
    
    def adv(self, start: int = 0, end: int = 0, path: Path | str = Path(), file: str = ''):
        '''Calculate pearson correlation coefficient of tickets rate 
        in `start` ~ `end` days before departure date'''
        
        filterwarnings("ignore")
        if not len(self.__merge):
            self.__merge = self.merge()
        data = self.__merge[self.__merge['day_adv'].isin(
            tuple(range(start if start > 0 else 1, end + 1 if end > 0 else self.__merge['day_adv'].max() + 1))
        )].reset_index()
        
        wb = self.indexbook(airlines = True)
        rules = (
            self.__diffrule('red', 'cellIs', '>', [0.9]), 
            self.__diffrule('yellow', 'cellIs', '>', [0.8]), 
            self.__diffrule('', 'cellIs', '>', [0.5]), 
            self.__diffrule('blue', 'cellIs', '<', [-0.9]), 
            self.__diffrule('green', 'cellIs', '<', [-0.8]), 
            self.__diffrule('', 'cellIs', '<', [-0.5]))
        percent, idct = -1, 0
        total = len(data) * 2
        data['date_flight'] = data['date_flight'].map(date.fromordinal)
        data['date_coll'] = data['date_coll'].map(date.fromordinal)
        
        overview = wb.create_sheet('总览')
        ov_title = sorted(data['airline'].unique(), reverse = True, 
            key = lambda x: len(self.__merge.loc[self.__merge['airline'] == x]))
        ov_index = sorted(data['route'].unique(), reverse = True, 
            key = lambda x: len(self.__merge.loc[self.__merge['route'] == x]))
        overview.cell(1, 1, '航线').font = self.set_bold
        overview.column_dimensions['A'].width = 13
        overview.freeze_panes = 'B2'
        for idx in range(len(ov_title)):
            cell = overview.cell(1, idx + 2, ov_title[idx])
            cell.font = self.set_hyperlink
            cell.hyperlink = f"#'{ov_title[idx]}'!A1"
        for idx in range(len(ov_index)):
            cell = overview.cell(idx + 2, 1, ov_index[idx])
            cell.font = self.set_hyperlink
            cell.hyperlink = f"#'{ov_index[idx]}'!A1"
        
        for airline, airlines in data.groupby(['airline']):
            idct += len(airlines)
            if percent != int(idct / total * 100):
                percent = int(idct / total * 100)
                print(f"\radv corr >> {percent:03d}", end = '%')
            col = 2 + ov_title.index(airline)
            ws = wb.create_sheet(airline)
            routes = sorted(airlines['route'].unique(), reverse = True, 
                key = lambda x: len(airlines.loc[airlines['route'] == x]))
            ws.append(['航班日期', '相关平均', '强正相关', '正相关', '负相关', '强负相关'] + routes)
            ws.auto_filter.ref = 'A1:F1'
            for idx in range(ws.max_column):
                cell = ws.cell(1, idx + 1)
                cell.alignment = self.set_align
                if idx >= 6:
                    cell.hyperlink = f"#'{cell.value}'!A1"
                    cell.font = self.set_hyperlink
                else:
                    cell.font = self.set_bold
            ws.freeze_panes = 'G2'
            for flt, flts in airlines.groupby(['date_flight']):
                if len(flts) <= 1:
                    continue
                corr = flts['price_rate'].corr(flts['day_adv'], 'pearson')
                output = [flt, corr, 0, 0, 0, 0] + list(None for _ in range(len(routes)))
                for route, group in flts.groupby(['route']):
                    corr = group['price_rate'].corr(group['day_adv'], 'pearson')
                    output[routes.index(route) + 6] = round(corr, 4)
                    if corr > 0.8:
                        output[2] += 1
                    elif corr > 0:
                        output[3] += 1
                    elif corr < -0.8:
                        output[5] += 1
                    elif corr < 0:
                        output[4] += 1
                ws.append(output)
                ws.cell(ws.max_row, 1).number_format = self.set_date
            rstring = f"G2:{ws.cell(ws.max_row, ws.max_column).coordinate}"
            for rule in rules:
                ws.conditional_formatting.add(rstring, rule)
            ws.conditional_formatting.add(
                f"C2:C{ws.max_row}", self.__diffrule(0, 'cellIs', '>=', [1]))
            ws.conditional_formatting.add(
                f"F2:F{ws.max_row}", self.__diffrule(0, 'cellIs', '>=', [1]))
            
            cell = ws.cell(1, 1)
            cell.hyperlink = f"#'{self.index_name}'!A1"
            cell.font = self.set_hyperlink
            cell = ws.cell(ws.max_row + 1, 1, '返回索引')
            cell.hyperlink = f"#'{self.index_name}'!A1"
            cell.font = self.set_hyperlink
            for idx in range(7, ws.max_column + 1):
                cell = ws.cell(ws.max_row, idx)
                coord = cell.coordinate
                _col = coord.strip(str(ws.max_row))
                cell.value = f'=AVERAGE({_col}2:{_col}{ws.max_row - 1})'
                overview.cell(2 + ov_index.index(ws.cell(1, idx).value), col).value = \
                    f'={airline}!{coord}'
        
        rstring = f"B2:{overview.cell(overview.max_row, overview.max_column).coordinate}"
        for rule in rules:
            overview.conditional_formatting.add(rstring, rule)
        
        for route, routes in data.groupby(['route']):
            idct += len(routes)
            if percent != int(idct / total * 100):
                percent = int(idct / total * 100)
                print(f"\radv corr >> {percent:03d}", end = '%')
            ws = wb.create_sheet(route)
            airlines = sorted(routes['airline'].unique(), reverse = True, 
                key = lambda x: len(routes.loc[routes['airline'] == x]))
            ws.append(['航班日期', '相关平均', '强正相关', '正相关', '负相关', '强负相关'] + airlines)
            ws.auto_filter.ref = 'A1:F1'
            for idx in range(ws.max_column):
                cell = ws.cell(1, idx + 1)
                cell.alignment = self.set_align
                if idx >= 6:
                    cell.hyperlink = f"#'{cell.value}'!A1"
                    cell.font = self.set_hyperlink
                else:
                    cell.font = self.set_bold
            ws.freeze_panes = 'G2'
            for flt, flts in routes.groupby(['date_flight']):
                if len(flts) <= 1:
                    continue
                corr = flts['price_rate'].corr(flts['day_adv'], 'pearson')
                output = [flt, corr, 0, 0, 0, 0] + list(None for _ in range(len(airlines)))
                for airline, group in flts.groupby(['airline']):
                    corr = group['price_rate'].corr(group['day_adv'], 'pearson')
                    output[airlines.index(airline) + 6] = round(corr, 4)
                    if corr > 0.8:
                        output[2] += 1
                    elif corr > 0:
                        output[3] += 1
                    elif corr < -0.8:
                        output[5] += 1
                    elif corr < 0:
                        output[4] += 1
                ws.append(output)
                ws.cell(ws.max_row, 1).number_format = self.set_date
            rstring = f"{ws.cell(2, 7).coordinate}:{ws.cell(ws.max_row, ws.max_column).coordinate}"
            for rule in rules:
                ws.conditional_formatting.add(rstring, rule)
            ws.conditional_formatting.add(
                f"C2:C{ws.max_row}", self.__diffrule(0, 'cellIs', '>=', [1]))
            ws.conditional_formatting.add(
                f"F2:F{ws.max_row}", self.__diffrule(0, 'cellIs', '>=', [1]))
            
            cell = ws.cell(1, 1)
            cell.hyperlink = f"#'{self.index_name}'!A1"
            cell.font = self.set_hyperlink
            cell = ws.cell(ws.max_row + 1, 1, '返回索引')
            cell.hyperlink = f"#'{self.index_name}'!A1"
            cell.font = self.set_hyperlink
            cell = ws.cell(ws.max_row, 2, '返程')
            route = route.split('-', 1)
            cell.hyperlink = f"#'{route[1]}-{route[0]}'!A1"
            cell.font = self.set_hyperlink
            for idx in range(7, ws.max_column + 1):
                cell = ws.cell(ws.max_row, idx)
                coord = cell.coordinate
                col = coord.strip(str(ws.max_row))
                cell.value = f'=AVERAGE({col}2:{col}{ws.max_row - 1})'
        
        print("\rsaving   ")
        filterwarnings("default")
        if file == '' or file is None:
            file = f"corr_adv_{self.__root.name}.xlsx"
        elif not file.endswith(".xlsx"):
            file += ".xlsx"
        if not isinstance(path, Path):
            path = Path(path)
        path.mkdir(parents = True, exist_ok = True)
        if (path / Path(file)).exists():
            time = datetime.today().strftime("%H%M%S")
            file = file.replace(".xlsx", f"_{time}.xlsx")
        wb.save(path / Path(file))
        wb.close()
    
    
    def week(self, *days: int, limit: int = 5, \
        path: Path | str = Path(), file: str = ''):
        '''Calculate pearson correlation coefficient of `days` of week'''
        filterwarnings("ignore")
        if not len(self.__merge):
            self.__merge = self.merge()
        data = self.__merge.drop(self.__merge[self.__merge['day_adv'] < limit].index)
        data['month'] = data['date_flight'].map(lambda x: date.fromordinal(x).month)
        for month in data['month'].unique():
            months = data.loc[data['month'] == month]
            if months['date_flight'].nunique() < 7:
                data.drop(months.index, inplace = True)
        data['day_week'] = data['day_week'].map({
            '星期一': 1, '星期二': 2, '星期三': 3, '星期四': 4, 
            '星期五': 5, '星期六': 6, '星期日': 7})
        
        if not len(days):
            days = [1, 2, 3, 4, 5, 6, 7]
        wb = Workbook()
        rules = (
            self.__diffrule('red', 'cellIs', '>', [0.5]), 
            self.__diffrule('yellow', 'cellIs', '>', [0.3]), 
            self.__diffrule('blue', 'cellIs', '<', [-0.5]), 
            self.__diffrule('green', 'cellIs', '<', [-0.3]))
        percent, idct = -1, 0
        total = len(data) * len(days)
        airlines = sorted(data['airline'].unique(), reverse = True, 
            key = lambda x: len(data.loc[data['airline'] == x]))
        
        for day in days:
            ws = wb.create_sheet(str(day))
            ws.append(['出发', '到达', '相关平均', '强正相关', '正相关', '负相关', '强负相关'] + airlines)
            ws.auto_filter.ref = 'A1:G1'
            for idx in range(ws.max_column):
                ws.cell(1, idx + 1).font = self.set_bold
                ws.cell(1, idx + 1).alignment = self.set_align
            ws.freeze_panes = 'H2'
            data['target'] = data['day_week'].map(lambda x: 1 if x == day else 0)
            for route, routes in data.groupby(['route']):
                idct += len(routes)
                if percent != int(idct / total * 100):
                    percent = int(idct / total * 100)
                    print(f"\rweek corr >> {percent:03d}", end = '%')
                corr = mean(list(months['price_rate'].corr(months['target'], 'pearson') \
                    for _, months in routes.groupby(['month'])))
                if corr == nan or len(routes.loc[routes['day_week'] == day]) < \
                    routes['date_coll'].nunique() * routes['date_flight'].nunique() / 2:
                    continue
                output = route.split('-', 1) + [corr, 0, 0, 0, 0] + \
                    list(None for _ in range(len(airlines)))
                for airline, group in routes.groupby(['airline']):
                    if len(group.loc[group['day_week'] == day]) <= limit:
                        continue
                    corr = mean(list(months['price_rate'].corr(months['target'], 'pearson') \
                        for _, months in group.groupby(['month'])))
                    output[airlines.index(airline) + 7] = round(corr, 4)
                    if corr > 0.3:
                        output[3] += 1
                    elif corr > 0:
                        output[4] += 1
                    elif corr < -0.3:
                        output[6] += 1
                    elif corr < 0:
                        output[5] += 1
                ws.append(output)
            for idx in range(1, ws.max_row):
                if -0.1 <= ws.cell(idx + 1, 3).value <= 0.1:
                    ws.row_dimensions[idx + 1].hidden = 1
            rstring = f"H2:{ws.cell(ws.max_row, ws.max_column).coordinate}"
            for rule in rules:
                ws.conditional_formatting.add(rstring, rule)
            ws.conditional_formatting.add(
                f"D2:D{ws.max_row}", self.__diffrule(0, 'cellIs', '>=', [1]))
            ws.conditional_formatting.add(
                f"G2:G{ws.max_row}", self.__diffrule(0, 'cellIs', '>=', [1]))
        
        print("\rsaving   ")
        filterwarnings("default")
        if file == '' or file is None:
            file = f"corr_week_{self.__root.name}.xlsx"
        elif not file.endswith(".xlsx"):
            file += ".xlsx"
        if not isinstance(path, Path):
            path = Path(path)
        path.mkdir(parents = True, exist_ok = True)
        if (path / Path(file)).exists():
            time = datetime.today().strftime("%H%M%S")
            file = file.replace(".xlsx", f"_{time}.xlsx")
        wb.remove(wb.active)
        wb.save(path / Path(file))
        wb.close()
    
    
    def hour(self, key: Literal['day_week', 'airline'], start: int = 6, end: int = 24,
             limit: int = 5, path: Path | str = Path(), file: str = ''):
        '''Calculate pearson correlation coefficient of hours 
        from `start` to `end` in group of `key`'''
        filterwarnings("ignore")
        if not len(self.__merge):
            self.__merge = self.merge()
        data = self.__merge.drop(self.__merge[self.__merge['day_adv'] < limit].index)
        data['date_flight'] = data['date_flight'].map(date.fromordinal)
        
        wb = Workbook()
        rules = (
            self.__diffrule('red', 'cellIs', '>', [0.7]), 
            self.__diffrule('yellow', 'cellIs', '>', [0.5]), 
            self.__diffrule('', 'cellIs', '>', [0.3]), 
            self.__diffrule('blue', 'cellIs', '<', [-0.7]), 
            self.__diffrule('green', 'cellIs', '<', [-0.5]), 
            self.__diffrule('', 'cellIs', '<', [-0.3]))
        percent, idct = -1, 0
        hours = tuple(range(start, end))
        total = len(data) * len(hours)
        for hour in hours:
            ws = wb.create_sheet(str(hour))
            targets = sorted(
                data.loc[data['hour_dep'] == hour][key].unique(), \
                    key = lambda x: len(data.loc[data[key] == x]), reverse = True) \
                if key == 'airline' else \
                    ['星期一', '星期二', '星期三', '星期四', '星期五', '星期六', '星期日'] \
                if key == 'day_week' else \
                    sorted(data.loc[data['hour_dep'] == hour][key].unique())
            ws.append(['出发', '到达', '相关平均', '强正相关', '正相关', '负相关', '强负相关'] + targets)
            ws.auto_filter.ref = 'A1:G1'
            for idx in range(ws.max_column):
                ws.cell(1, idx + 1).font = self.set_bold
                ws.cell(1, idx + 1).alignment = self.set_align
            ws.freeze_panes = 'H2'
            data['target'] = data['hour_dep'].map(lambda x: 1 if x == hour else 0)
            for route, routes in data.groupby(['route']):
                idct += len(routes)
                if percent != int(idct / total * 100):
                    percent = int(idct / total * 100)
                    print(f"\rhour corr >> {percent:03d}", end = '%')
                corr = mean(list(colls['price_rate'].corr(colls['target'], 'pearson') \
                    for _, colls in routes.groupby(['date_coll'])))
                if corr == nan or len(routes.loc[routes['hour_dep'] == hour]) < \
                    routes['date_coll'].nunique() * routes['date_flight'].nunique() / 2:
                    continue
                output = route.split('-', 1) + [corr, 0, 0, 0, 0] + \
                    list(None for _ in range(len(targets)))
                for item, group in routes.groupby([key]):
                    if len(group.loc[group['hour_dep'] == hour]) <= limit:
                        continue
                    corr = mean(list(colls['price_rate'].corr(colls['target'], 'pearson') \
                        for _, colls in group.groupby(['date_coll'])))
                    output[targets.index(item) + 7] = round(corr, 4)
                    if corr > 0.5:
                        output[3] += 1
                    elif corr > 0:
                        output[4] += 1
                    elif corr < -0.5:
                        output[6] += 1
                    elif corr < 0:
                        output[5] += 1
                ws.append(output)
            for idx in range(1, ws.max_row):
                value = ws.cell(idx + 1, 3).value
                if -0.1 <= value <= 0.1:
                    ws.row_dimensions[idx + 1].hidden = 1
            rstring = f"H2:{ws.cell(ws.max_row, ws.max_column).coordinate}"
            for rule in rules:
                ws.conditional_formatting.add(rstring, rule)
            ws.conditional_formatting.add(
                f"D2:D{ws.max_row}", self.__diffrule(0, 'cellIs', '>=', [1]))
            ws.conditional_formatting.add(
                f"G2:G{ws.max_row}", self.__diffrule(0, 'cellIs', '>=', [1]))
        
        print("\rsaving   ")
        filterwarnings("default")
        if file == '' or file is None:
            file = f"corr_hour_{key}_{self.__root.name}.xlsx"
        elif not file.endswith(".xlsx"):
            file += ".xlsx"
        if not isinstance(path, Path):
            path = Path(path)
        path.mkdir(parents = True, exist_ok = True)
        if (path / Path(file)).exists():
            time = datetime.today().strftime("%H%M%S")
            file = file.replace(".xlsx", f"_{time}.xlsx")
        wb.remove(wb.active)
        wb.save(path / Path(file))
        wb.close()
