import datetime
from datetime import date, time
import json
import re
import openpyxl
from openpyxl.styles import Font, Alignment
import pathlib
import asyncio
import aiohttp
import ssl
def ignore(codeList:list) -> dict:
    '''If two cities given are too close or not in analysis range, skip, 
    by key-ing the coordinates and value-ing False of the coordinate in dict. '''
    codesum=len(codeList)
    ignoreDict=codeDict={}
    ignoreDict=ignoreDict.fromkeys(ignoreList,True)
    # Every city tuple above is designated to be ignored, therefore their values are True.
    for i in range(codesum):
        for j in range(i,codesum):
            if i==j:
                continue
            if ignoreDict.get((codeList[j],codeList[i]),False) or ignoreDict.get((codeList[i],codeList[j]),False):
                # If the city tuple is not found in the dict, it won't have a value.
                # If the city tuple is found in the dict, its value is False since it shouldn't be processed.
                codeDict[(i,j)]=False
                codeDict[(j,i)]=False
    #print(codeDict)
    return codeDict
async def getProxy() -> str:
    try:    # Set proxy: run in Docker / cmd -> cd ProxyPool -> docker-compose up -> (idle)
        async with aiohttp.ClientSession() as newProxy:
            async with newProxy.get('http://127.0.0.1:5555/random') as proxy:
                proxy = await proxy.text()
            proxy = proxy.strip()
    except:
        for i in range(5):
            try:
                async with aiohttp.ClientSession() as newProxy:
                    async with newProxy.get('http://127.0.0.1:5555/random') as proxy:
                        proxy = await proxy.text()
                    proxy = proxy.strip()
                    break
            except:
                continue
        return ''
    return "http://"+proxy  #should be str in aiohttp
async def getWebpage(session:aiohttp.ClientSession, url: str, data: str, header: dict, proxy: str) -> list:
    FORCED_CIPHERS = (
    'ECDH+AESGCM:DH+AESGCM:ECDH+AES256:DH+AES256:ECDH+AES128:DH+AES:ECDH+HIGH:'
    'DH+HIGH:ECDH+3DES:DH+3DES:RSA+AESGCM:RSA+AES:RSA+HIGH:RSA+3DES')
    sslcontext = ssl.create_default_context()
    sslcontext.set_ciphers(FORCED_CIPHERS)
    try:
        sslcontext.options |= ssl.OP_NO_SSLv2
        async with session.post(url, data=data, headers=header, proxy=proxy, ssl=sslcontext) as reply:
            response = await reply.text()
            #print(response)
            routeList = json.loads(response).get('data').get('routeList')   # -> list
            #print(routeList)
            return routeList
    except:
        pass
    try:
        sslcontext.options |= ssl.OP_NO_SSLv3
        async with session.post(url, data=data, headers=header, proxy=proxy, ssl=sslcontext) as reply:
            response = await reply.text()
            #print(response)
            routeList = json.loads(response).get('data').get('routeList')   # -> list
            #print(routeList)
            return routeList
    except:
        pass
    try:
        async with session.post(url, data=data, headers=header, proxy=proxy, ssl=ssl._create_unverified_context) as reply:
            response = await reply.text()
            #print(response)
            routeList = json.loads(response).get('data').get('routeList')   # -> list
            #print(routeList)
            return routeList
    except:
        pass
    return list()
async def getTickets(session: aiohttp.ClientSession, fdate: date, dcity: str, acity: str) -> int:
    "Get tickets from dep city to arr city on one date, put all data to the excel, and return a status code."
    global idct
    dcityname = airportCity.get(dcity, None)
    acityname = airportCity.get(acity, None)
    dow=dayOfWeek[fdate.isoweekday()]
    datarows = []
    url = "https://flights.ctrip.com/itinerary/api/12808/products"
    header = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:68.0) Gecko/20100101 Firefox/68.0",
        "Referer": "https://flights.ctrip.com/itinerary/oneway/BJS-SHA?date="+fdate.isoformat(),
        "Content-Type": "application/json"}
    payload = {
        "flightWay": "Oneway",
        "classType": "ALL",
        "hasChild": False,
        "hasBaby": False,
        "searchIndex": 1,
        "airportParams": [
            {"dcity": dcity, "acity": acity, "dcityname": dcityname, "acityname": acityname,
                "date": fdate.isoformat(), "dcityid": 1, "acityid": 2}]}
    proxy = await getProxy()
    if proxy == '':
        idct+=1
        return 4    #代理异常
    routeList = await getWebpage(session,url,json.dumps(payload),header,proxy)
    if routeList == []:
        for i in range(5):
            try:
                proxy = await getProxy()
                routeList = await getWebpage(session,url,json.dumps(payload),header,proxy)
                if routeList != []:
                    break
            except:
                continue
        idct+=1
        return 2    #网络或源数据异常
    try:
        for route in routeList:
            if len(route.get('legs')) == 1:
                legs = route.get('legs')
                #print(legs,end='\n\n')
                flight = legs[0].get('flight')
                if flight.get('sharedFlightNumber'):
                    continue    # Shared flights not collected
                airlineName = flight.get('airlineName')
                if '旗下' in airlineName:   # Airline name should be as simple as possible
                    airlineName = airlineName.split('旗下',1)[1]
                departureTime = flight.get('departureDate').split(' ',1)[1].split(':',2)
                departureTime = time(int(departureTime[0]),int(departureTime[1]))   # Convert the time string to a time class
                arrivalTime = flight.get('arrivalDate').split(' ',1)[1].split(':',2)
                arrivalTime = time(int(arrivalTime[0]),int(arrivalTime[1])) # Convert the time string to a time class
                if dcityname == '北京' or dcityname == '上海' or dcityname== '成都':
                    departureName = flight.get('departureAirportInfo').get('airportName')
                    departureName = dcityname + re.match('成?都?(.*?)国?际?机场',departureName).groups()[0]
                elif dcityname: # If dcityname exists, that means the code-name is in the default code-name dict
                    departureName = dcityname   # Multi-airport cities need the airport name while others do not
                else:   # If dcityname is None, that means the code-name is not in the default code-name dict
                    departureName = flight.get('departureAirportInfo').get('cityName')
                    airportCity[dcity] = departureName  # ...therefore it is added now 
                if acityname == '北京' or acityname == '上海' or acityname== '成都':
                    arrivalName = flight.get('arrivalAirportInfo').get('airportName')
                    arrivalName = acityname + re.match('成?都?(.*?)国?际?机场',arrivalName).groups()[0]
                elif acityname:
                    arrivalName = acityname
                else:
                    arrivalName = flight.get('arrivalAirportInfo').get('cityName')
                    airportCity[acity] = arrivalName
                craftType = flight.get('craftTypeKindDisplayName')
                craftType = re.match('(.)型',craftType).groups()[0] # Crafttype is S/M/L
                ticket = legs[0].get('cabins')[0]   # Price info in cabins dict
                price = ticket.get('price').get('price')
                rate = ticket.get('price').get('rate')
                datarows.append([fdate,dow,airlineName,craftType,departureName,arrivalName,departureTime,arrivalTime,price,rate,])
                #日期，星期，航司，机型，出发机场，到达机场，出发时间，到达时间，价格，折扣
    except:
        idct+=1
        return 3    #数据处理异常
    if len(datarows)<=3:    # Data count too less, ignore these flights in the future.
        idct+=1
        return 1    #数据过少并忽略
    wbook = openpyxl.Workbook()
    wsheet=wbook.active
    for row in datarows:
        wsheet.append(row)
    wbook.save(corrDate+'\\{0}_{1}-{2}.xlsx'.format(fdate.isoformat(),dcity,acity))
    wbook.close
    print('\r',end='')
    for i in range(round(20*idct/total)):
        print('>',end='')
    for i in range(round(20*idct/total),20):
        print('-',end='')
    idct+=1
    return 0    #数据已抓取
async def generateXlsx(fdate:date,days:int,codeList:list):
    "Generate excels of flight tickets info, between the citys given and from the date given and days needed."
    cdate=fdate
    global dayOfWeek, airportCity, idct, total
    dayOfWeek={1:'星期一',2:'星期二',3:'星期三',4:'星期四',5:'星期五',6:'星期六',7:'星期日'}
    airportCity={
        'BJS':'北京','CAN':'广州','SHA':'上海','CTU':'成都','TFU':'成都','SZX':'深圳','KMG':'昆明','XIY':'西安','PEK':'北京',
        'PKX':'北京','PVG':'上海','CKG':'重庆','HGH':'杭州','NKG':'南京','CGO':'郑州','XMN':'厦门','WUH':'武汉','CSX':'长沙',
        'TAO':'青岛','HAK':'海口','URC':'乌鲁木齐','TSN':'天津','KWE':'贵阳','HRB':'哈尔滨','SHE':'沈阳','SYX':'三亚','DLC':'大连',
        'TNA':'济南','NNG':'南宁','LHW':'兰州','FOC':'福州','TYN':'太原','CGQ':'长春','KHN':'南昌','HET':'呼和浩特','NGB':'宁波',
        'WNZ':'温州','ZUH':'珠海','HFE':'合肥','SJW':'石家庄','INC':'银川','YNT':'烟台','KWL':'桂林','JJN':'泉州','WUX':'无锡',
        'SWA':'揭阳','XNN':'西宁','LJG':'丽江','JHG':'西双版纳','NAY':'北京','LXA':'拉萨','MIG':'绵阳','CZX':'常州','NTG':'南通',
        'YIH':'宜昌','WEH':'威海','XUZ':'徐州','ZHA':'湛江','YTY':'扬州','DYG':'张家界','DSN':'鄂尔多斯','BHY':'北海','LYI':'临沂',
        'HLD':'呼伦贝尔','HUZ':'惠州','UYN':'榆林','YCU':'运城','KHG':'喀什','HIA':'淮安','BAV':'包头','ZYI':'遵义','KRL':'库尔勒',
        'LUM':'德宏','YNZ':'盐城','KOW':'赣州','YIW':'义乌','LYG':'连云港','XFN':'襄阳','CIF':'赤峰','LZO':'泸州','DLU':'大理',
        'AKU':'阿克苏','YNJ':'延吉','ZYI':'遵义','HTN':'和田','LZH':'柳州','LYA':'洛阳','WDS':'十堰','HSN':'舟山','JNG':'济宁',
        'YIN':'伊宁','ENH':'恩施','ACX':'兴义','HYN':'台州','TCZ':'腾冲','DAT':'大同','BSD':'保山','BFJ':'毕节','NNY':'南阳',
        'WXN':'万州','TGO':'通辽','CGD':'常德','HNY':'衡阳','XIC':'西昌','MDG':'牡丹江','RIZ':'日照','NAO':'南充','YBP':'宜宾',
    }
    checkdict=ignore(codeList)  # The keys in the dict are the coordinates that should not be processed, whose values are False.
    codesum=len(codeList)
    idct=0
    total=codesum*(codesum-1)-(len(checkdict)/2)
    tasks=[]
    print('处理中: ')
    for dcity in range(codesum):
        for acity in range(dcity,codesum):
            if acity == dcity:  # Same values should not be processed.
                continue
            if checkdict.get((dcity,acity),True): # If the city tuple key / coordinate is not found, process.
                async with aiohttp.ClientSession(connector=aiohttp.TCPConnector(limit=64, loop=loop)) as session:
                    for i in range(days):
                        tasks.append(asyncio.create_task(getTickets(session, cdate, codeList[dcity], codeList[acity])))
                        tasks.append(asyncio.create_task(getTickets(session, cdate, codeList[acity], codeList[dcity])))
                        cdate=cdate.fromordinal(cdate.toordinal()+1)    #one day forward
                    await asyncio.wait(tasks)
                tasks=[]
                cdate=fdate #reset when reaching the last day
            else:
                break
def arrangeXlsx(path:str):
    "Format every excel file and return the number of formatted files."
    i=0
    for file in pathlib.Path(path).iterdir():
        if file.match('*.xlsx') and '~' not in file.name:   # Skip temp (hidden) excels
            wbook=openpyxl.load_workbook(file)
            wsheet=wbook.active
            if wsheet['A1'].value == '日期':    # Sign of formatted
                wbook.close
                continue
            i+=1
            wsheet.insert_rows(1,1)
            wsheet['A1'].value = '日期'
            wsheet['B1'].value = '星期'
            wsheet['C1'].value = '航司'
            wsheet['D1'].value = '机型'
            wsheet['E1'].value = '出发机场'
            wsheet['F1'].value = '到达机场'
            wsheet['G1'].value = '出发时'
            wsheet['H1'].value = '到达时'
            wsheet['I1'].value = '价格'
            wsheet['J1'].value = '折扣'
            wsheet.column_dimensions['A'].width = 11
            wsheet.column_dimensions['B'].width = 7
            wsheet.column_dimensions['C'].width = 12
            wsheet.column_dimensions['D'].width = 6
            wsheet.column_dimensions['G'].width = 7.5
            wsheet.column_dimensions['H'].width = 7.5
            wsheet.column_dimensions['I'].width = 6
            wsheet.column_dimensions['J'].width = 6
            for row in wsheet.iter_rows(1,1,1,10):
                for cell in row:
                    cell.alignment=Alignment(vertical='center',horizontal='center')
                    cell.font=Font(bold=True)
            for row in wsheet.iter_rows(2,wsheet.max_row,7,8):  # Adjust the dep time and arr time formats
                for cell in row:
                    cell.number_format='HH:MM'
            for row in wsheet.iter_rows(2,wsheet.max_row,10,10):    # Make the rate show as percentage
                for cell in row:
                    cell.number_format='0%'
            for row in wsheet.iter_rows(2,wsheet.max_row,4,8):  # Alignment adjusts
                for cell in row:
                    cell.alignment=Alignment(vertical='center',horizontal='center')
            wbook.save(path+'\\'+file.name)
            wbook.close
    return i
if __name__ == "__main__":  #务必先设置代理：win+R -> cmd -> cd ProxyPool -> docker-compose up -> (idle) -> start
    global corrDate, ignoreList
    corrDate = str(datetime.datetime.now().date())
    path=pathlib.Path(corrDate)
    if not path.exists():
        pathlib.Path.mkdir(path)
    ignoreList=[('BJS','TSN'),('BJS','SJW'),('BJS','TYN'),('BJS','TNA'),('BJS','SHE'),('BJS','HET'),
                ('SJW','TYN'),('SJW','TNA'),('TSN','TNA'),('TSN','TYN'),('TYN','TNA'),('SJW','TSN'),
                ('SHE','CGQ'),('CGQ','HRB'),('SHE','HRB'),('DLC','SHE'),('DLC','CGQ'),('TSN','DLC'),
                ('BJS','SHE'),('BJS','CGO'),('TNA','CGO'),('BJS','TAO'),('TNA','TAO'),('TSN','TAO'),
                ('CGO','SJW'),('CGO','XIY'),('CGO','HFE'),('CGO','TYN'),('CGO','WUH'),('CGO','NKG'),
                ('XIY','INC'),('XIY','LHW'),('CTU','XIY'),('XIY','TYN'),('XIY','SJW'),('XIY','WUH'),
                ('WUH','HFE'),('WUH','NKG'),('NKG','HFE'),('WUH','CSX'),('WUH','HGH'),('WUH','KHN'),
                ('NKG','WUX'),('NKG','CZX'),('NKG','NTG'),('NKG','YTY'),('NKG','SHA'),('NKG','HGH'),
                ('SHA','HGH'),('SHA','WUX'),('SHA','NTG'),('SHA','CZX'),('SHA','YTY'),('HGH','WUX'),
                ('HGH','CZX'),('HGH','NTG'),('HGH','YTY'),('WUX','CZX'),('WUX','NTG'),('WUX','YTY'),
                ('CZX','NTG'),('CZX','YTY'),('NTG','YTY'),('SHA','HFE'),('HGH','HFE'),('HFE','CZX'),
                ('HFE','WUX'),('HFE','YTY'),('HFE','NTG'),('WUH','CZX'),('WUH','WUX'),('WUH','NTG'),
                ('WUH','YTY'),('KHN','CSX'),('KHN','HGH'),('KHN','FOC'),('KHN','XMN'),('KHN','KWE'),
                ('CSX','CAN'),('CSX','NNG'),('CSX','FOC'),('CSX','XMN'),('HGH','FOC'),('HGH','XMN'),
                ('CAN','SZX'),('CAN','ZUH'),('ZUH','SZX'),('CAN','SWA'),('ZUH','SWA'),('SZX','SWA'),
                ('SWA','FOC'),('SWA','XMN'),('CAN','NNG'),('FOC','NNG'),('KWE','NNG'),('KWE','KMG'),
                ('KMG','JHG'),('KWE','CTU'),('KWE','CSX'),('CKG','KWE'),('CTU','CKG'),('CKG','XIY'),
                ('LHW','XNN'),('XNN','INC'),('INC','LHW'),('HET','INC'),('HET','TYN'),('HET','SJW'),]
    #可手动添加不参与爬取的城市对
    loop = asyncio.get_event_loop()
    loop.run_until_complete(generateXlsx(date(2022,2,1),2,['BJS','CAN','ZHA',]))
    #调参得表: 起始年月日、往后天数、机场三字码列表
    print('共整理',arrangeXlsx(corrDate),'个文件')  #整理表格