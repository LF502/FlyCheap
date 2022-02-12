import pandas
import pathlib
import openpyxl
__airports = {'北京': 1, '上海': 1, '广州': 1, 
              '成都': 0.8, '深圳': 0.75, '昆明': 0.7, '西安': 0.65, 
              '重庆': 0.65, '杭州': 0.6, '南京': 0.45, '郑州': 0.4, '厦门': 0.4, 
              '武汉': 0.4, '长沙': 0.4, '青岛': 0.4, '海口': 0.35, '乌鲁木齐': 0.35, 
              '天津': 0.35, '贵阳': 0.3, '哈尔滨': 0.3, '沈阳': 0.3, '三亚': 0.3, 
              '大连': 0.3, '济南': 0.25, '南宁': 0.25, '兰州': 0.2, '福州': 0.2, 
              '太原': 0.2, '长春': 0.2, '南昌': 0.2, '呼和浩特': 0.2, '宁波': 0.2, 
              '温州': 0.2, '珠海': 0.2, '合肥': 0.2, '石家庄': 0.15, '银川': 0.15, 
              '烟台': 0.15, '桂林': 0.1, '泉州': 0.1, '无锡': 0.1, '揭阳': 0.1, 
              '西宁': 0.1, '丽江': 0.1, '西双版纳': 0.1, '南阳': 0.1,}
__airportCity = {'BJS':'北京','CAN':'广州','SHA':'上海','CTU':'成都','TFU':'成都','SZX':'深圳','KMG':'昆明','XIY':'西安','PEK':'北京',
                 'PKX':'北京','PVG':'上海','CKG':'重庆','HGH':'杭州','NKG':'南京','CGO':'郑州','XMN':'厦门','WUH':'武汉','CSX':'长沙',
                 'TAO':'青岛','HAK':'海口','URC':'乌鲁木齐','TSN':'天津','KWE':'贵阳','HRB':'哈尔滨','SHE':'沈阳','SYX':'三亚','DLC':'大连',
                 'TNA':'济南','NNG':'南宁','LHW':'兰州','FOC':'福州','TYN':'太原','CGQ':'长春','KHN':'南昌','HET':'呼和浩特','NGB':'宁波',
                 'WNZ':'温州','ZUH':'珠海','HFE':'合肥','SJW':'石家庄','INC':'银川','YNT':'烟台','KWL':'桂林','JJN':'泉州','WUX':'无锡',
                 'SWA':'揭阳','XNN':'西宁','LJG':'丽江','JHG':'西双版纳','NAY':'北京','LXA':'拉萨','MIG':'绵阳','CZX':'常州','NTG':'南通',
                 'YIH':'宜昌','WEH':'威海','XUZ':'徐州','ZHA':'湛江','YTY':'扬州','DYG':'张家界','DSN':'鄂尔多斯','BHY':'北海','LYI':'临沂',
                 'HLD':'呼伦贝尔','HUZ':'惠州','UYN':'榆林','YCU':'运城','KHG':'喀什','HIA':'淮安','BAV':'包头','ZYI':'遵义','KRL':'库尔勒',
                 'LUM':'德宏','YNZ':'盐城','KOW':'赣州','YIW':'义乌','LYG':'连云港','XFN':'襄阳','CIF':'赤峰','LZO':'泸州','DLU':'大理',
                 'AKU':'阿克苏','YNJ':'延吉','ZYI':'遵义','HTN':'和田','LZH':'柳州','LYA':'洛阳','WDS':'十堰','HSN':'舟山','JNG':'济宁',
                 'YIN':'伊宁','ENH':'恩施','ACX':'兴义','HYN':'台州','TCZ':'腾冲','DAT':'大同','BSD':'保山','BFJ':'毕节','NNY':'南阳',
                 'WXN':'万州','TGO':'通辽','CGD':'常德','HNY':'衡阳','XIC':'西昌','MDG':'牡丹江','RIZ':'日照','NAO':'南充','YBP':'宜宾',}
def run(path: pathlib.Path = pathlib.Path()):
    wb = openpyxl.Workbook()
    sheets = ["总表", "干线", "小干线", "支线",]
    for sheet in sheets:
        wb.create_sheet(sheet)
    wb.remove(wb.active)
    airlinelist = []
    rows = {"干线": [], "小干线": [], "支线": []}
    for file in path.iterdir():
        if not file.match('*.xlsx'):
            continue
        print('\r' + file.name, end = ' processing...')
        data = pandas.read_excel(file.joinpath()).iloc[ : , [0, 2, 9]]
        name = file.name.split('~')
        dcity = __airportCity.get(name[0])
        acity = __airportCity.get(name[1].strip('_preproc.xlsx'))
        name = dcity + ' - ' + acity
        airlinedict = {}
        for item in data.values:
            if item[1] not in airlinelist:
                airlinelist.append(item[1])
            if not airlinedict.get(item[0]):
                airlinedict[item[0]] = {item[1]: []}
            elif not airlinedict.get(item[0]).get(item[1]):
                airlinedict[item[0]][item[1]] = []
            airlinedict[item[0]][item[1]].append(item[2])
        del data
        avgdict =  {"平均": []}
        for airline in airlinelist:
            avgdict[airline] = []
        for day in airlinedict.keys():
            total = rate = 0
            sheet = day.date().isoformat()
            if sheet not in wb.sheetnames:
                wb.create_sheet(sheet)
                rows[sheet] = []
            for airlines in airlinedict[day].values():
                total += len(airlines)
                for value in airlines:
                    rate += value
            rate /= total
            avgdict["平均"].append(rate)
            row = [name, rate, ]
            for airline in airlinelist:
                if airlinedict[day].get(airline):
                    item = 0
                    for value in airlinedict[day][airline]:
                        item += value
                    avg = item / len(airlinedict[day][airline]) / rate
                    avgdict[airline].append(avg)
                    row.append(avg)
                else:
                    row.append(None)
            rows[sheet].append(row)
        row = [name, ]
        for value in avgdict.values():
            if len(value):
                avg = 0
                for item in value:
                    avg += item
                avg /= len(value)
                row.append(avg)
            else:
                row.append(None)
        apf = __airports.get(dcity, 0.05) + __airports.get(acity, 0.05)
        if apf >= 1.4:
            rows["干线"].append(row)
        elif apf > 1:
            rows["小干线"].append(row)
        else:
            rows["支线"].append(row)
    wb["总表"].append(["城市对", "平均折扣"] + airlinelist)
    for sheet in rows.keys():
        wb[sheet].append(["城市对", "平均折扣"] + airlinelist)
        for row in rows[sheet]:
            if sheet in sheets:
                wb["总表"].append(row)
            wb[sheet].append(row)
    for sheet in wb:
        sheet.column_dimensions["A"].width = 16
        sheet.freeze_panes = 'B2'
        sheet.append(("平均", ))
        for col in range(2, sheet.max_column + 1):
            coordinate = sheet.cell(sheet.max_row, col).coordinate
            sheet[coordinate] = f"=AVERAGE({sheet.cell(2, col).coordinate}:{sheet.cell(sheet.max_row - 1, col).coordinate})"
    wb.save(f"airline-rate_{path.name}.xlsx")
    wb.close
    print('\nDone!')
        
if __name__ == "__main__":
    run(pathlib.Path('2022-02-09'))